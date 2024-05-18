import re
import subprocess
from multiprocessing import Process, Manager
import os
from scapy.all import Ether, IP, TCP, UDP, sendpfast, Raw
import time
from tqdm import tqdm
from openpyxl import Workbook
import pickle

def get_cases(pkt_types, pkt_lengths, cap_ppses):
    cases = []
    for pkt_type in pkt_types:
        for pkt_length in pkt_lengths:
            for cap_pps in cap_ppses:
                cases.append((pkt_type, pkt_length, cap_pps))
    return cases

def pad_message(m, l, protocol="TCP"):
    TCP_LENGTH = 54
    UDP_LENGTH = 42
    active_length = TCP_LENGTH if protocol == "TCP" else UDP_LENGTH
    padding_message = "-PADDING-"
    lm = len(m)
    padded_len = l - (active_length + lm)
    padding_message *= int(1 + padded_len / len(padding_message))
    padded_message = m + padding_message
    return padded_message[:(l - active_length)]

def execute_command(command, return_dict):
    try:
        process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        output, error = process.communicate()
        return_dict["output"] = output
        return_dict["error"] = error
        return_dict["return_code"] = process.returncode
    except subprocess.CalledProcessError as e:
        return_dict["output"] = None
        return_dict["error"] = str(e)
        return_dict["return_code"] = e.returncode

def get_packet(preset, length, protocol):
    basepkt = Ether(src="08:00:27:00:00:02", dst="08:00:27:00:00:01") / IP(src="10.147.18.200", dst="192.168.1.56")
    message = pad_message(f"{preset.capitalize()} Packet", length, protocol)
    if protocol == "TCP":
        if preset == "harmless":
            return basepkt / TCP(sport=5000, dport=1234) / Raw(load=message)
        elif preset == "suspicious":
            return basepkt / TCP(sport=5000, dport=80) / Raw(load=message)
        elif preset == "harmful":
            return basepkt / TCP(sport=5000, dport=22) / Raw(load=message)
        elif preset == "mixed":
            message2 = pad_message("Suspicious Packet", length + 10, protocol)
            return [
                basepkt / TCP(sport=5000, dport=1234) / Raw(load=message),
                basepkt / TCP(sport=5000, dport=80) / Raw(load=message2)
            ]
    else:
        if preset == "harmless":
            return basepkt / UDP(dport=1234) / Raw(load=message)
        elif preset == "suspicious":
            return basepkt / UDP(dport=80) / Raw(load=message)
        elif preset == "harmful":
            return basepkt / UDP(dport=22) / Raw(load=message)
        elif preset == "mixed":
            message2 = pad_message("Suspicious Packet", length + 10, protocol)
            return [
                basepkt / UDP(dport=1234) / Raw(load=message),
                basepkt / UDP(dport=80) / Raw(load=message2)
            ]

def execute_test(pkt_type, pkt_length, cap_pps, injector, sniffer, pkt_repeat, timeout, active_protocol):
    command = ["/home/jkulrativid/Desktop/Piglet/softwares/gateway/perf-sniffer", sniffer, "\"\"", str(pkt_length), str(pkt_repeat), "0", "0", str(timeout)]
    if pkt_type == "mixed":
        command = ["/home/jkulrativid/Desktop/Piglet/softwares/gateway/perf-sniffer", sniffer, "\"\"", str(pkt_length), str(pkt_repeat), str(pkt_length + 10), str(pkt_repeat), str(timeout)]

    manager = Manager()
    return_dict = manager.dict()
    p = Process(target=execute_command, args=(command, return_dict))
    p.start()

    time.sleep(1)

    pkt = get_packet(pkt_type, pkt_length, active_protocol)
    sendpfast(pkt, iface=injector, loop=pkt_repeat, file_cache=True, mbps=1000, pps=cap_pps)

    p.join()

    output, error, return_code = return_dict["output"], return_dict["error"], return_dict["return_code"]

    if return_code != 0:
        print("Error occurred while executing the command.")
        print("Error message:")
        print(error)
        exit(1)

    duration = re.search(r"Duration: (.+?) s, (.+?) ns", output).group(1)
    throughput_mbps = re.search(r"Throughput\(Mbps\): (.+?) Mbps", output).group(1)
    pps = re.search(r"Packet per second: (.+?) pps", output).group(1)
    pkt_count = re.search(r"overall pkt1 count = ([0-9]+)", output).group(1)
    if pkt_type == "mixed":
        pkt_count2 = re.search(r"overall pkt2 count = ([0-9]+)", output).group(1)
        pkt_count = f"{int(pkt_count) + int(pkt_count2)} (pkt1: {pkt_count}, pkt2: {pkt_count2})"
    print(f"Duration: {duration} s")
    print(f"Throughput: {throughput_mbps} Mbps")
    print(f"Packet per second: {pps} pps")
    print(f"Packet count: {pkt_count}")
    return duration, throughput_mbps, pps, pkt_count

def write_single_run(ws, base_row, base_col, result):
    pkt_type, pkt_length, cap_pps, duration, throughput_mbps, pps, pkt_count = result
    ws.cell(row=base_row, column=base_col, value=throughput_mbps)
    ws.cell(row=base_row + 1, column=base_col, value=pps)
    ws.cell(row=base_row + 2, column=base_col, value=duration)
    ws.cell(row=base_row + 3, column=base_col, value=pkt_count)

def save_to_excel(results, dir, title):
    row_increment = 4
    col_increment = 3
    wb = Workbook()
    ws = wb.active
    ws.title = title
    row = 1
    col = 1
    for i, runs in enumerate(results):
        for j, result in enumerate(runs):
            write_single_run(ws, row, col + j, result)
        row += row_increment
        if row % (4 * row_increment) == 1:
            col += col_increment
            row = 1
    wb.save(f"{dir}/{title}.xlsx")

def main():
    injector = "enx2887ba3e44aa" # tp-link white
    sniffer = "enxc84d442973a0" # gray

    save_to_dir = "results-sat-18"

    pkt_types = ["harmless", "suspicious", "harmful", "mixed"]
    pkt_lengths = [65, 100, 500, 1500]
    cap_ppses = [5000, 10000, 20000, 24000]
    # pkt_repeat = 200_000 if test_title != "mixed" else 100_000
    timeout = 50
    protocols = ["TCP", "UDP"]
    for active_protocol in protocols:
        for pkt_type in pkt_types:
            test_title = f"{pkt_type}_{active_protocol}"
            pkt_repeat = 200_000 if pkt_type != "mixed" else 100_000
            # pkt_repeat = 2

            print(f"testing {test_title} packets")
            print(f"pkt_types: {pkt_types}")
            print(f"pkt_lengths: {pkt_lengths}")
            print(f"cap_ppses: {cap_ppses}")
            print(f"pkt_repeat: {pkt_repeat}")
            print(f"timeout: {timeout}")
            print(f"active_protocol: {active_protocol}")

            print("injector:", injector)
            print("sniffer:", sniffer)

            cases = get_cases([pkt_type], pkt_lengths, cap_ppses)
            results = []
            for case in tqdm(cases):
                pkt_type, pkt_length, cap_pps = case
                print(f"Executing test case: {pkt_type}, {pkt_length}, {cap_pps}\r")
                runs = []
                for i in range(3):
                    duration, throughput_mbps, pps, pkt_count = execute_test(pkt_type, pkt_length, cap_pps, injector, sniffer, pkt_repeat, timeout, active_protocol)
                    runs.append((pkt_type, pkt_length, cap_pps, duration, throughput_mbps, pps, pkt_count))
                results.append(runs)

            if not os.path.exists(save_to_dir):
                os.makedirs(save_to_dir)
            if not os.path.exists(f"{save_to_dir}/{active_protocol}"):
                os.makedirs(f"{save_to_dir}/{active_protocol}")

            save_to_excel(results, f"{save_to_dir}/{active_protocol}", test_title)
            with open(f"{save_to_dir}/{active_protocol}/{test_title}.pkl", "wb") as f:
                pickle.dump(results, f)

if __name__ == '__main__':
    main()
